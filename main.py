import os
import sys
import argparse
import getpass
import shutil
import tempfile
from tqdm import tqdm
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import NamedStyle, Protection
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError as e:
    print(f"An error occurred while importing necessary modules: {e}. Please make sure you have the openpyxl module installed. You can install dependencies by running 'pip install -r requirements.txt'")
    sys.exit(1)

def unlock_cells_and_dropdowns(input_file=None, output_file=None, sheet_name=None, password=None, progress_bar=True):
    """
    Unlocks all cells and dropdowns in a specified Excel sheet and applies password protection.
    
    Parameters:
    input_file (str, optional): The path to the input Excel file. If None, the user will be prompted to enter it.
    output_file (str, optional): The path to the output Excel file. If None, the user will be prompted to enter it.
    sheet_name (str, optional): The name of the sheet to unlock and protect. If None, the user will be prompted to enter it.
    password (str, optional): The password to apply to the sheet. If None, the user will be prompted to enter a password.
    progress_bar (bool, optional): Whether to display a progress bar during processing. Default is False.
    
    Returns:
    None
    
    Raises:
    ValueError: If the sheet does not exist in the workbook or if the password confirmation fails after 3 attempts.
    PermissionError: If the program does not have permission to save the file to the specified output path.
    """
    if input_file is None:
        input_file = input("Please enter the path to the input Excel file: ")
    if output_file is None:
        output_file = input("Please enter the path to the output Excel file: ")
    if sheet_name is None:
        sheet_name = input("Please enter the name of the sheet to unlock and protect: ")

    attempts = 0
    print("Starting process...")
    # Check that the input file exists and is a valid Excel file.
    if not os.path.isfile(input_file):
        raise FileNotFoundError(f"The input file {input_file} does not exist. Please check your file path.")
    try:
        workbook = load_workbook(input_file, read_only=False, keep_vba=True, data_only=True, keep_links=False)
    except (InvalidFileException, Exception) as e:
        raise ValueError(f"An error occurred while opening the input file: {e}. Please check if the input file is a valid Excel file and that it is not corrupted.")

    # Check that the output directory exists and create it if necessary.
    output_dir = os.path.dirname(output_file)
    if not os.path.isdir(output_dir):
        os.makedirs(output_dir)

    # Make a copy of the workbook to avoid modifying the original.
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, os.path.basename(input_file))
    shutil.copyfile(input_file, temp_file)
    workbook = load_workbook(temp_file)

    # Verify that the sheet exists in the workbook.
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"The sheet {sheet_name} does not exist in the workbook. Please check your sheet name.")
    sheet = workbook[sheet_name]

    # Create a new style or retrieve an existing one.
    if "unlocked_style" in workbook.named_styles:
        unlocked_style = workbook.named_styles["unlocked_style"]
    else:
        unlocked_style = NamedStyle(name="unlocked_style")
        unlocked_style.protection = Protection(locked=False)
        workbook.add_named_style(unlocked_style)

    # Apply the style to all cells.
    if progress_bar:
        total_cells = sheet.max_row * sheet.max_column
        progress = tqdm(total=total_cells)
        for row in sheet.iter_rows():
            for cell in row:
                cell.style = unlocked_style
                progress.update()

    # Modify dropdown settings.
    print("Modifying dropdown settings...")
    if sheet.data_validations.dataValidation:
        for dv in sheet.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1:
                if dv.formula1.startswith("="):
                    named_range = dv.formula1[1:]
                    if named_range in workbook.defined_names.definedName:
                        dv.formula1 = workbook.defined_names.definedName[named_range].attr_text
        print("Dropdown settings modified successfully.")
    else:
        print("No dropdown settings found.")

    # Apply password protection.
    if password is None:
        while attempts < 3:
            password = getpass.getpass(prompt="Please enter the password to protect the sheet: ")
            confirm_password = getpass.getpass(prompt="Please confirm your password: ")
            if password == confirm_password:
                break
            else:
                print("Passwords do not match. Please try again.")
            attempts += 1
        if attempts == 3:
            raise ValueError("Failed to confirm password after 3 attempts. Please make sure you're entering the same password for confirmation.")

    sheet.protection.set_password(password)
    print("Password protection applied successfully.")

    # Save the modified workbook to the output file.
    print(f"Saving workbook to {output_file}...")
    try:
        workbook.save(output_file)
    except PermissionError as e:
        raise PermissionError(f"Permission denied when trying to save the file to {output_file}. Error message: {e}")

    print("Workbook saved successfully. Process complete.")
    shutil.rmtree(temp_dir)

    print(f"Process complete. The sheet {sheet_name} has been unlocked and saved to {output_file}.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Unlock cells and dropdowns in an Excel sheet.")
    parser.add_argument("--input_file", help="The path to the input Excel file.")
    parser.add_argument("--output_file", help="The path to the output Excel file.")
    parser.add_argument("--sheet_name", help="The name of the sheet to unlock and protect.")
    parser.add_argument("--password", help="The password to apply to the sheet.")
    parser.add_argument("--progress_bar", action="store_true", help="Whether to display a progress bar during processing.")
    args = parser.parse_args()

    unlock_cells_and_dropdowns(args.input_file, args.output_file, args.sheet_name, args.password, args.progress_bar)
